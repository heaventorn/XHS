# -*- coding: utf-8 -*-
"""core/storage.py — 双层 EXCEL 落盘（公共核心模块）

框架升级：把输出拆成两个 Sheet，对应两条采集通道——
  Sheet1「线索池」：社交平台（小红书/知乎/B站）采到的帖子/账号，联系方式留空待人工补
  Sheet2「联系池」：工商数据（企查查/天眼查）直接拿到的公司电话/邮箱

两个 Sheet 都以「公司名」为关联键，可在 Excel 里 VLOOKUP 打通。
支持实时落盘（采到一条立即写入），中途退出不丢已采集数据。
"""
import os
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


# ---------------- Sheet1 线索池：社交平台 ----------------
CLUE_HEADERS = [
    "序号", "来源平台", "搜索关键词", "内容标题", "内容链接",
    "作者昵称", "作者ID", "角色推断", "关联公司", "匹配板块",
    "联系方式", "跟进状态", "采集时间",
]
CLUE_COL_WIDTHS = [6, 12, 14, 34, 46, 18, 16, 12, 18, 12, 18, 14, 20]

# ---------------- Sheet2 联系池：工商数据 ----------------
CONTACT_HEADERS = [
    "序号", "公司名", "统一社会信用代码", "所属板块", "融资轮次",
    "法定代表人/董监高", "联系电话", "公司邮箱", "注册地址",
    "公司官网", "融资新闻链接", "联系人角色", "跟进状态", "采集时间",
]
CONTACT_COL_WIDTHS = [6, 24, 22, 12, 10, 22, 16, 26, 30, 28, 40, 14, 14, 20]


def _style_sheet(ws, headers, widths, header_fill_hex):
    """应用表头样式与列宽（公共）。"""
    header_font = Font(name="微软雅黑", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color=header_fill_hex, end_color=header_fill_hex, fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9"),
    )
    for col_idx, (header, width) in enumerate(zip(headers, widths), 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border
        ws.column_dimensions[cell.column_letter].width = width

    data_font = Font(name="微软雅黑", size=10)
    data_align = Alignment(horizontal="left", vertical="top", wrap_text=True)
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.font = data_font
            cell.alignment = data_align
            cell.border = thin_border
    ws.row_dimensions.height = 30
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def _write_rows(ws, headers, rows, offset):
    """从指定行号开始写入数据行（offset 为已有行数，用于增量写）。"""
    for i, item in enumerate(rows):
        row_idx = offset + i + 2  # 表头占1行
        values = [i + 1] + [item.get(h, "") for h in headers[1:]]
        for col_idx, val in enumerate(values, 1):
            ws.cell(row=row_idx, column=col_idx, value=val)


def save_two_sheets(filepath, clue_rows=None, contact_rows=None, live=False):
    """保存双层 EXCEL。filepath 为输出路径；clue_rows/contact_rows 为各自数据列表。

    实时模式（live=True）时打开已有文件增量追加，避免覆盖已采数据。
    """
    clue_rows = clue_rows or []
    contact_rows = contact_rows or []

    if live and os.path.exists(filepath):
        # 增量模式：打开现有文件，追加新行
        try:
            from openpyxl import load_workbook
            wb = load_workbook(filepath)
            if "线索池" in wb.sheetnames and "联系池" in wb.sheetnames:
                ws_clue = wb["线索池"]
                ws_contact = wb["联系池"]
                clue_offset = ws_clue.max_row - 1
                contact_offset = ws_contact.max_row - 1
                _write_rows(ws_clue, CLUE_HEADERS, clue_rows, clue_offset)
                _write_rows(ws_contact, CONTACT_HEADERS, contact_rows, contact_offset)
                wb.save(filepath)
                return filepath
        except Exception:
            pass  # 文件损坏等异常 → 走全量重建

    # 全量模式：重建两个 Sheet
    wb = Workbook()
    ws_clue = wb.active
    ws_clue.title = "线索池"
    _style_sheet(ws_clue, CLUE_HEADERS, CLUE_COL_WIDTHS, "FF2442")  # 线索池红头
    _write_rows(ws_clue, CLUE_HEADERS, clue_rows, 0)

    ws_contact = wb.create_sheet("联系池")
    _style_sheet(ws_contact, CONTACT_HEADERS, CONTACT_COL_WIDTHS, "1F9D55")  # 联系池绿头
    _write_rows(ws_contact, CONTACT_HEADERS, contact_rows, 0)

    wb.save(filepath)
    return filepath


def default_output(prefix="采集结果"):
    """生成默认输出文件名（时间戳）。"""
    return os.path.join(os.path.expanduser("~"), "Desktop",
                        f"{prefix}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")
