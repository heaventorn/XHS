# -*- coding: utf-8 -*-
"""
小红书关键词帖子采集脚本（反检测加固版 v2.0.6）
功能：搜索指定关键词，采集帖子的账号名、帖子标题、帖子链接，输出到 Excel
      采到一条立即写入「实时结果文件」，中途退出/崩溃也不丢已采集数据
反爬强化：偏态人类延迟、真实硬件值注入、关键词乱序+采集量浮动、代理/总量配置
          plugins/mimeTypes 运行时透传真实值、permissions.toString 伪装 native code
v2.0.6：详情页/搜索兜底跳转注入 Referer（模拟点击链）、拦截 favicon 消除双峰时序
使用：python xhs_keyword_scraper.py
依赖：pip install playwright openpyxl
"""

import os
import re
import sys
import time
import random
import math
import json
from datetime import datetime, timedelta
from urllib.parse import urljoin, quote
from playwright.sync_api import sync_playwright
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# ==================== 配置区 ====================
KEYWORDS = [
    "FA",
    "寻找FA",
    "一级市场",
    "融资顾问",
    "FA合作",
    "FA建联",
    "FA交流",
    "找FA",
    "一级市场融资",
    "FA服务",
    "FA业务",
]

MAX_PER_KEYWORD = 10
FILTER_WITHIN_TWO_YEARS = True
OUTPUT_FILE = os.path.join(os.path.expanduser("~"), "Desktop", f"小红书关键词采集_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")
# 实时结果文件（固定名）：每采到 1 条就增量写入，中途退出/崩溃也能保住已采集数据
LIVE_OUTPUT_FILE = os.path.join(os.path.expanduser("~"), "Desktop", "小红书关键词采集_实时.xlsx")
MAX_SCROLLS = 4

# 代理设置：留空 = 直连本机 IP；填入后走代理（如 socks5://127.0.0.1:7890 或 http://127.0.0.1:7890）
# 提示：单 IP 高频是触发风控的首要原因，条件允许建议使用固定住宅代理
PROXY_URL = ""

# 单次运行总采集上限：防止单 IP 一次性抓太多触发频控（达到后提前结束）
MAX_TOTAL_NOTES = 60

# 账号黑名单：这些账号发布的帖子会被跳过，不采集（精确匹配，不区分大小写）
BLACKLIST_AUTHORS = [
    # "账号昵称1",
    # "账号昵称2",
]

# 账号小红书号黑名单：这些小红书号（纯数字ID）发布的帖子会被跳过，不采集
# 与 BLACKLIST_AUTHORS（昵称）二选一或同时使用均可
BLACKLIST_IDS = [
    "109181692",
    "49292951005",
]

# 内容不相关关键词黑名单：标题包含这些词的帖子会被跳过（不区分大小写）
# 用于剔除搜索结果中与 FA/融资/一级市场无关的内容，可根据实际情况增删
IRRELEVANT_KEYWORDS = [
    # 养生健康类
    "养生", "保健", "中医", "中药", "艾灸", "拔罐", "刮痧", "推拿", "按摩", "针灸",
    "减肥", "瘦身", "健身", "瑜伽", "冥想",
    # 相亲情感类
    "相亲", "恋爱", "情感", "婚姻", "择偶", "脱单", "约会", "表白", "分手", "复合",
    "渣男", "渣女", "绿茶", "海王", "舔狗", "备胎", "暧昧", "暗恋", "初恋", "异地恋",
    # 美妆时尚类
    "美妆", "护肤", "化妆", "美甲", "美睫", "美容", "美发", "染发", "烫发", "穿搭",
    "时尚", "潮流", "奢侈品", "包包", "首饰",
    # 美食类
    "美食", "料理", "食谱", "烹饪", "烘焙", "甜品", "蛋糕", "火锅", "烧烤", "奶茶",
    "咖啡", "茶叶", "红酒", "白酒", "啤酒",
    # 旅游生活类
    "旅游", "旅行", "攻略", "景点", "打卡", "拍照", "摄影", "vlog", "日常", "生活",
    "好物", "推荐", "种草", "测评", "开箱",
    # 母婴教育类
    "母婴", "育儿", "亲子", "孕妇", "宝宝", "婴儿", "教育", "学习", "考试", "考研",
    "考公", "留学", "英语",
    # 职场副业类
    "招聘", "求职", "面试", "简历", "职场", "副业", "兼职", "赚钱", "打工", "辞职",
    "离职", "跳槽",
    # 房产汽车类
    "买房", "租房", "装修", "家居", "家具", "家电", "汽车", "车型", "二手车",
    # 娱乐游戏类
    "游戏", "电竞", "动漫", "漫画", "小说", "影视", "电影", "电视剧", "音乐", "歌曲",
    "明星", "八卦", "娱乐", "综艺", "直播", "带货",
    # 玄学星座类
    "星座", "塔罗", "命理", "风水", "算命", "占卜", "玄学", "灵异", "鬼故事",
    # 宠物收藏类
    "宠物", "猫狗", "花鸟", "收藏", "古董", "文玩",
]


# 持久化浏览器用户目录：用于保存登录态，避免每次运行重新扫码
PROFILE_DIR = os.path.join(os.path.expanduser("~"), ".xhs_scraper_profile")

# 全量累积结果（跨关键词），用于实时落盘与中断时保存
_all_rows = []
_seen_note_ids = set()

# 卡片内常见发布时间元素（启发式；定位不到时回退到正则从卡片文本提取）
TIME_SELECTORS = [
    "span[class*='date']",
    "span[class*='time']",
    "time",
    "div[class*='date']",
]

# ==================== 反检测配置 ====================
# 兜底 User-Agent（仅读取真实 UA 失败时使用）；正常运行时动态跟随真实浏览器版本
USER_AGENT = "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/138.0.0.0 Safari/537.36"
VIEWPORT = {"width": 1920, "height": 1080}
LANGUAGES = ["zh-CN", "zh", "en-US", "en"]

# 分层延迟（秒）
DELAY_PAGE_LOAD = (10.0, 15.0)
DELAY_AFTER_SEARCH = (8.0, 12.0)
DELAY_AFTER_SCROLL = (5.0, 8.0)
DELAY_BETWEEN_KEYWORDS = (30.0, 60.0)
DELAY_MOUSE_MOVE = (0.6, 1.4)
DELAY_EXTRACT = (0.5, 1.2)
DELAY_READ_PAUSE = (3.0, 6.0)  # 阅读停顿

# ========== 深度 Stealth 注入脚本（v1 优化版） ==========
STEALTH_SCRIPT_TEMPLATE = r"""
// ===== 全局常量 =====
const __UA = "__UA_PLACEHOLDER__";

// 基于 seed 的伪随机数生成器（保证同一 canvas 多次调用结果一致，避免完全随机被检测）
function __seededRandom(seed) {
    let s = seed % 2147483647;
    if (s <= 0) s += 2147483646;
    return function() {
        s = (s * 16807) % 2147483647;
        return (s - 1) / 2147483646;
    };
}

// ===== 1. 核心自动化特征掩盖（含原型链）=====
const __wdDesc = { get: () => false, set: () => {}, configurable: true, enumerable: true };
Object.defineProperty(navigator, 'webdriver', __wdDesc);
try { Object.defineProperty(Navigator.prototype, 'webdriver', __wdDesc); } catch(e) {}
Object.defineProperty(navigator, 'automation', { get: () => undefined, configurable: true });
Object.defineProperty(navigator, 'controlledByAutomat', { get: () => undefined, configurable: true });

// 原型链注入痕迹清理
if (navigator.__proto__) {
    const __np = navigator.__proto__;
    ['__driver_evaluate','__webdriver_evaluate','__selenium_evaluate','__fxdriver_evaluate',
     '__driver_unwrapped','__webdriver_unwrapped','__selenium_unwrapped','__fxdriver_unwrapped',
     '__webdriver_script_fn','__webdriver_script_func','__webdriver_script_timeout',
     '_cdc_adoQpoasnfa76pfcZLmcfl_','_cdc_asdjflasutopfhvcZLmcfl_','_cdc_sdjflasutopfhvcZLmcfl_']
    .forEach(function(p){ try{ delete __np[p]; }catch(e){} });
}

// ===== 2. Navigator 完整属性伪造（plugins/mimeTypes 运行时透传真实值，避免数量/内容与真机不符）=====
const __plugins = __PLUGINS_JSON__;
__plugins.item = function(i){ return __plugins[i] || null; };
__plugins.namedItem = function(n){ return __plugins.find(function(p){return p.name===n;}) || null; };
__plugins.refresh = function(){};
Object.defineProperty(navigator, 'plugins', { get: function(){ return __plugins; }, configurable: true });

const __mimes = __MIMES_JSON__;
__mimes.item = function(i){ return __mimes[i] || null; };
__mimes.namedItem = function(n){ return __mimes.find(function(m){return m.type===n;}) || null; };
Object.defineProperty(navigator, 'mimeTypes', { get: function(){ return __mimes; }, configurable: true });

const __navProps = {
    languages: ['zh-CN','zh','en-US','en'],
    language: 'zh-CN',
    hardwareConcurrency: __HW_CONCURRENCY__,
    deviceMemory: __DEVICE_MEMORY__,
    maxTouchPoints: __MAX_TOUCH_POINTS__,
    vendor: '__VENDOR__',
    platform: '__PLATFORM__',
    cookieEnabled: true,
    onLine: true,
    doNotTrack: null,
    vendorSub: '',
    productSub: '20030107',
    product: 'Gecko',
    appVersion: __UA.replace('Mozilla/',''),
    userAgent: __UA,
    appName: 'Netscape',
    appCodeName: 'Mozilla',
};
Object.keys(__navProps).forEach(function(key){
    try { Object.defineProperty(navigator, key, { get: function(){ return __navProps[key]; }, configurable: true }); } catch(e){}
});

// ===== 2.5 userAgentData（UA Client Hints）伪造：与 navigator.userAgent 版本对齐 =====
try {
    if (!("userAgentData" in navigator)) {
        var __brands = [
            {brand: "Chromium", version: "__CHROME_MAJOR_PLACEHOLDER__"},
            {brand: "Google Chrome", version: "__CHROME_MAJOR_PLACEHOLDER__"},
            {brand: "Not.A/Brand", version: "24"}
        ];
        var __uaData = {
            brands: __brands,
            mobile: false,
            platform: "Windows",
            getHighEntropyValues: function() {
                return Promise.resolve({
                    architecture: "x86",
                    bitness: "64",
                    brands: __brands,
                    fullVersionList: [
                        {brand: "Chromium", version: "__CHROME_VERSION_PLACEHOLDER__"},
                        {brand: "Google Chrome", version: "__CHROME_VERSION_PLACEHOLDER__"},
                        {brand: "Not.A/Brand", version: "24.0.0.0"}
                    ],
                    mobile: false,
                    model: "",
                    platform: "Windows",
                    platformVersion: "15.0.0",
                    uaFullVersion: "__CHROME_VERSION_PLACEHOLDER__",
                    wow64: false
                });
            },
            toJSON: function() { return {brands: __brands, mobile: false, platform: "Windows"}; }
        };
        Object.defineProperty(navigator, "userAgentData", { get: function(){ return __uaData; }, configurable: true });
    }
} catch(e){}

// ===== 3. Screen 与窗口尺寸伪造 =====
const __screenProps = { width:1920, height:1080, availWidth:1920, availHeight:1040, availLeft:0, availTop:0, colorDepth:24, pixelDepth:24 };
Object.keys(__screenProps).forEach(function(key){
    try { Object.defineProperty(screen, key, { get: function(){ return __screenProps[key]; }, configurable: true }); } catch(e){}
});
try {
    Object.defineProperty(window, 'outerWidth', { get: function(){ return 1920; }, configurable: true });
    Object.defineProperty(window, 'outerHeight', { get: function(){ return 1080; }, configurable: true });
    Object.defineProperty(window, 'innerWidth', { get: function(){ return 1920; }, configurable: true });
    Object.defineProperty(window, 'innerHeight', { get: function(){ return 1040; }, configurable: true });
    Object.defineProperty(window, 'devicePixelRatio', { get: function(){ return 1; }, configurable: true });
} catch(e){}

// ===== 4. Chrome / Runtime 完整伪造 =====
window.chrome = window.chrome || {};
window.chrome.runtime = window.chrome.runtime || { id:'', onMessage:{}, sendMessage:function(){}, onConnect:{}, connect:function(){}, onInstalled:{}, onStartup:{}, connectNative:function(){} };
window.chrome.app = window.chrome.app || { isInstalled:false, getDetails:function(){}, getIsInstalled:function(){return false;}, runningState:'cannot run' };
window.chrome.csi = window.chrome.csi || function(){ return { startE:Date.now(), onloadT:Date.now(), pageT:100, tran:15 }; };
window.chrome.loadTimes = window.chrome.loadTimes || function(){ return {
    requestTime:Date.now()/1000, startLoadTime:Date.now()/1000, commitLoadTime:Date.now()/1000,
    finishDocumentLoadTime:Date.now()/1000, finishLoadTime:Date.now()/1000, firstPaintTime:Date.now()/1000,
    connectionInfo:'h2', npnNegotiatedProtocol:'h2', wasAlternateProtocolAvailable:false, wasFetchedViaSpdy:true, wasNpnNegotiated:true
};};

// ===== 5. Permissions 完整伪造 =====
if (navigator.permissions && navigator.permissions.query) {
    const __origQuery = navigator.permissions.query.bind(navigator.permissions);
    const __query = function(parameters) {
        var name = parameters.name;
        if (name === 'notifications') return Promise.resolve({ state: Notification.permission, onchange:null });
        if (name === 'geolocation') return Promise.resolve({ state:'prompt', onchange:null });
        if (name === 'camera' || name === 'microphone') return Promise.resolve({ state:'prompt', onchange:null });
        if (name === 'midi') return Promise.resolve({ state:'prompt', onchange:null });
        return __origQuery(parameters);
    };
    // 保留原生函数名与 toString 特征，避免 toString() 暴露自定义函数体
    try { Object.defineProperty(__query, 'name', { value: 'query', configurable: true }); } catch(e){}
    try { window.__spoofedFns = window.__spoofedFns || []; window.__spoofedFns.push(__query); } catch(e){}
    navigator.permissions.query = __query;
}
try { Object.defineProperty(Notification, 'permission', { get: function(){ return 'default'; }, configurable: true }); } catch(e){}

// ===== 6. Canvas 指纹保护（基于 seed 的一致噪声，非完全随机）=====
const __origToDataURL = HTMLCanvasElement.prototype.toDataURL;
HTMLCanvasElement.prototype.toDataURL = function(type) {
    if (type === 'image/png' && this.width > 0 && this.height > 0) {
        try {
            var ctx = this.getContext('2d');
            if (ctx) {
                var imgData = ctx.getImageData(0, 0, this.width, this.height);
                var rand = __seededRandom(this.width * 31 + this.height * 17 + 7);
                for (var i = 0; i < imgData.data.length; i += 4) {
                    if (rand() < 0.008) {
                        var delta = rand() < 0.5 ? 1 : -1;
                        imgData.data[i] = Math.max(0, Math.min(255, imgData.data[i] + delta));
                    }
                }
                ctx.putImageData(imgData, 0, 0);
            }
        } catch(e){}
    }
    return __origToDataURL.apply(this, arguments);
};

const __origGetImageData = CanvasRenderingContext2D.prototype.getImageData;
CanvasRenderingContext2D.prototype.getImageData = function() {
    var result = __origGetImageData.apply(this, arguments);
    if (result && result.data && result.width > 0 && result.height > 0) {
        try {
            var rand = __seededRandom(result.width * 13 + result.height * 7 + Math.floor(arguments[0]||0) + Math.floor(arguments[1]||0) + 3);
            for (var i = 0; i < result.data.length; i += 4) {
                if (rand() < 0.004) {
                    result.data[i] = Math.max(0, Math.min(255, result.data[i] + 1));
                }
            }
        } catch(e){}
    }
    return result;
};

// ===== 7. WebGL 指纹伪造（vendor/renderer 返回常见显卡，避免无 GPU 的 SwiftShader 暴露）=====
try {
    var __glGetParam = WebGLRenderingContext.prototype.getParameter;
    WebGLRenderingContext.prototype.getParameter = function(param) {
        if (param === 37445) return "__WEBGL_VENDOR__";
        if (param === 37446) return "__WEBGL_RENDERER__";
        return __glGetParam.call(this, param);
    };
} catch(e){}
try {
    var __gl2GetParam = WebGL2RenderingContext.prototype.getParameter;
    WebGL2RenderingContext.prototype.getParameter = function(param) {
        if (param === 37445) return "__WEBGL_VENDOR__";
        if (param === 37446) return "__WEBGL_RENDERER__";
        return __gl2GetParam.call(this, param);
    };
} catch(e){}

// ===== 8. AudioContext 指纹保护（含 getByteTimeDomainData）=====
if (window.AudioContext || window.webkitAudioContext) {
    var __AC = window.AudioContext || window.webkitAudioContext;
    var __origCreateAnalyser = __AC.prototype.createAnalyser;
    __AC.prototype.createAnalyser = function() {
        var analyser = __origCreateAnalyser.apply(this, arguments);
        var __origGetByteFreq = analyser.getByteFrequencyData;
        analyser.getByteFrequencyData = function(array) {
            __origGetByteFreq.apply(this, arguments);
            var rand = __seededRandom(array.length);
            for (var i = 0; i < array.length; i++) {
                if (rand() < 0.025) array[i] = Math.max(0, Math.min(255, array[i] + (rand()<0.5?1:-1)));
            }
        };
        if (analyser.getByteTimeDomainData) {
            var __origGetByteTime = analyser.getByteTimeDomainData;
            analyser.getByteTimeDomainData = function(array) {
                __origGetByteTime.apply(this, arguments);
                var rand = __seededRandom(array.length + 99);
                for (var i = 0; i < array.length; i++) {
                    if (rand() < 0.015) array[i] = Math.max(0, Math.min(255, array[i] + (rand()<0.5?1:-1)));
                }
            };
        }
        return analyser;
    };
}

// ===== 9. toString 伪装（含 eval/setTimeout/setInterval）=====
const __origToString = Function.prototype.toString;
Function.prototype.toString = function() {
    try {
        // 对被脚本覆盖的函数，返回 native code 伪装，防止 toString() 暴露自定义函数体
        try { if (window.__spoofedFns && window.__spoofedFns.indexOf(this) !== -1) return 'function query() { [native code] }'; } catch(e){}
        var result = __origToString.call(this);
        if (result.indexOf('[native code]') !== -1) return result;
        return result;
    } catch(e) { return 'function () { [native code] }'; }
};
try {
    eval.toString = function(){ return 'function eval() { [native code] }'; };
    setTimeout.toString = function(){ return 'function setTimeout() { [native code] }'; };
    setInterval.toString = function(){ return 'function setInterval() { [native code] }'; };
    Function.prototype.toString.toString = function(){ return 'function toString() { [native code] }'; };
} catch(e){}

// ===== 10. 错误栈伪装（清理 playwright/selenium 痕迹）=====
Object.defineProperty(Error, 'stackTraceLimit', { value: 10, writable: true });
try {
    if (Error.prepareStackTrace) {
        var __origPrepare = Error.prepareStackTrace;
        Error.prepareStackTrace = function(error, stack) {
            var result = __origPrepare(error, stack);
            if (typeof result === 'string') {
                return result.replace(/playwright|puppeteer|selenium|webdriver|cdc_/gi, '');
            }
            return result;
        };
    }
} catch(e){}

// ===== 11. 连接属性伪造 =====
Object.defineProperty(navigator, 'connection', {
    get: function(){ return {
        effectiveType:'4g', rtt:50, downlink:10, saveData:false,
        onchange:null, addEventListener:function(){}, removeEventListener:function(){}, dispatchEvent:function(){return true;}
    };}, configurable: true
});

// ===== 12. Intl 时区一致性伪造 =====
try {
    var __origResolved = Intl.DateTimeFormat.prototype.resolvedOptions;
    Intl.DateTimeFormat.prototype.resolvedOptions = function() {
        var opts = __origResolved.call(this);
        opts.timeZone = 'Asia/Shanghai';
        opts.locale = 'zh-CN';
        return opts;
    };
} catch(e){}

// ===== 13. matchMedia 伪造（prefers-color-scheme = light）=====
try {
    var __origMatchMedia = window.matchMedia;
    window.matchMedia = function(query) {
        var result = __origMatchMedia.call(this, query);
        if (query.indexOf('prefers-color-scheme') !== -1) {
            try { Object.defineProperty(result, 'matches', { get: function(){ return query.indexOf('light') !== -1; }, configurable: true }); } catch(e){}
        }
        return result;
    };
} catch(e){}

// ===== 14. 清理 window/document 上的自动化注入痕迹 =====
['__playwright_playwright','__pw_manual','__PW_SELENIUM__','_phantom','callPhantom','_selenium',
 'webdriver','__driver_evaluate','__webdriver_evaluate','__selenium_evaluate','__fxdriver_evaluate',
 'cdc_adoQpoasnfa76pfcZLmcfl_','cdc_asdjflasutopfhvcZLmcfl_','cdc_sdjflasutopfhvcZLmcfl_']
.forEach(function(p){
    try { delete window[p]; } catch(e){}
    try { delete document[p]; } catch(e){}
});

// ===== 15. document.documentElement.lang =====
try { document.documentElement.lang = 'zh-CN'; } catch(e){}

// ===== 17. iframe 防护：监听新 iframe 并注入基础伪装 =====
try {
    var __iframeObserver = new MutationObserver(function(mutations) {
        mutations.forEach(function(mutation) {
            mutation.addedNodes.forEach(function(node) {
                if (node.tagName === 'IFRAME') {
                    try {
                        var idoc = node.contentDocument || (node.contentWindow && node.contentWindow.document);
                        if (idoc) idoc.documentElement.lang = 'zh-CN';
                    } catch(e){}
                }
            });
        });
    });
    __iframeObserver.observe(document.documentElement, { childList: true, subtree: true });
} catch(e){}

// ===== 18. 字体指纹伪造（document.fonts.status = loaded）=====
try {
    if (document.fonts) {
        Object.defineProperty(document.fonts, 'status', { get: function(){ return 'loaded'; }, configurable: true });
    }
} catch(e){}

// ===== 19. 电池 API 伪造（插电+满电，桌面端典型状态）=====
try {
    if ('getBattery' in navigator) {
        navigator.getBattery = function() {
            return Promise.resolve({
                charging: true, chargingTime: 0, dischargingTime: Infinity, level: 1.0,
                onchargingchange: null, onchargingtimechange: null, ondischargingtimechange: null, onlevelchange: null,
                addEventListener: function(){}, removeEventListener: function(){}, dispatchEvent: function(){ return true; }
            });
        };
    }
} catch(e){}

// ===== 20. WebGL debug extension 伪造（与 getParameter 伪造保持一致）=====
try {
    var __glExt = WebGLRenderingContext.prototype.getExtension;
    WebGLRenderingContext.prototype.getExtension = function(name) {
        if (name === 'WEBGL_debug_renderer_info') {
            return { UNMASKED_VENDOR_WEBGL: 37445, UNMASKED_RENDERER_WEBGL: 37446 };
        }
        return __glExt.call(this, name);
    };
} catch(e){}

"""


def build_stealth_script(real_ua, nav=None):
    """用真实 User-Agent 与真实硬件/平台值替换注入模板占位符，生成与运行环境一致的注入脚本。

    修复原硬编码 UA/硬件值与真实浏览器版本不一致、反而暴露自动化特征的问题；
    同时从 UA 解析出 Chrome 版本号，用于 userAgentData（UA Client Hints）伪造，
    保证 navigator.userAgent 与 navigator.userAgentData.brands 的版本完全一致。
    nav 为运行时读取的真实导航属性 dict，未提供时使用安全兜底值。
    """
    nav = nav or {}
    hw = int(nav.get("hw") or 8)
    mem = float(nav.get("mem") or 8)
    mtp = int(nav.get("mtp") or 0)
    platform = str(nav.get("platform") or "Win32")
    vendor = str(nav.get("vendor") or "Google Inc.")
    webgl_vendor = str(nav.get("webgl_vendor") or "Google Inc.")
    webgl_renderer = str(nav.get("webgl_renderer") or "ANGLE")
    # 运行时透传真实 plugins / mimeTypes（读取失败时回退到保守的常见值）
    plugins_json = json.dumps(nav.get("plugins") or [
        {"name": "PDF Viewer", "filename": "mhjfbmdgcfjbbpaeojofohoefgiehjai", "description": ""},
        {"name": "Chrome PDF Viewer", "filename": "mhjfbmdgcfjbbpaeojofohoefgiehjai", "description": ""},
        {"name": "Chromium PDF Viewer", "filename": "mhjfbmdgcfjbbpaeojofohoefgiehjai", "description": ""},
        {"name": "Microsoft Edge PDF Viewer", "filename": "mhjfbmdgcfjbbpaeojofohoefgiehjai", "description": ""},
        {"name": "WebKit built-in PDF", "filename": "internal-pdf-viewer", "description": "Portable Document Format"},
    ], ensure_ascii=False)
    mimes_json = json.dumps(nav.get("mimes") or [
        {"type": "application/pdf", "suffixes": "pdf", "description": ""},
        {"type": "text/pdf", "suffixes": "pdf", "description": ""},
    ], ensure_ascii=False)
    m = re.search(r"Chrome/(\d+)(?:\.(\d+)\.(\d+)\.(\d+))?", real_ua)
    if m:
        major = m.group(1)
        full_ver = f"{m.group(1)}.{m.group(2) or '0'}.{m.group(3) or '0'}.{m.group(4) or '0'}"
    else:
        major, full_ver = "138", "138.0.7204.0"
    return (STEALTH_SCRIPT_TEMPLATE
            .replace("__UA_PLACEHOLDER__", real_ua)
            .replace("__CHROME_MAJOR_PLACEHOLDER__", major)
            .replace("__CHROME_VERSION_PLACEHOLDER__", full_ver)
            .replace("__HW_CONCURRENCY__", str(hw))
            .replace("__DEVICE_MEMORY__", str(mem))
            .replace("__MAX_TOUCH_POINTS__", str(mtp))
            .replace("__PLATFORM__", platform)
            .replace("__VENDOR__", vendor)
            .replace("__WEBGL_VENDOR__", webgl_vendor)
            .replace("__WEBGL_RENDERER__", webgl_renderer)
            .replace("__PLUGINS_JSON__", plugins_json)
            .replace("__MIMES_JSON__", mimes_json))
# =================================================


def _skewed_random(lo, hi):
    """人类长尾偏态随机：多数时候接近下限，偶尔出现长停顿。
    真实人类操作间隔近似长尾/幂律分布，均匀分布反而像机器。"""
    if hi <= lo:
        return lo
    # random()**k 偏向小值，k 越大越偏下限；k 本身带随机，避免模式固定
    p = random.random() ** random.uniform(1.8, 3.2)
    return lo + (hi - lo) * p


def human_sleep(delay_range):
    time.sleep(_skewed_random(*delay_range))


def move_mouse_human(page, target_x=None, target_y=None, jitter=True):
    """人类化鼠标移动：贝塞尔曲线+随机抖动+速度变化+中途停顿+末端微抖动"""
    try:
        if target_x is None or target_y is None:
            target_x = random.randint(100, VIEWPORT["width"] - 100)
            target_y = random.randint(100, VIEWPORT["height"] - 200)

        start_x = random.randint(300, 900)
        start_y = random.randint(200, 700)

        steps = random.randint(25, 45)
        cp1_x = random.randint(0, VIEWPORT["width"])
        cp1_y = random.randint(0, VIEWPORT["height"])
        cp2_x = random.randint(0, VIEWPORT["width"])
        cp2_y = random.randint(0, VIEWPORT["height"])

        # 随机选择一个中途停顿点（15%概率）
        pause_at = random.uniform(0.3, 0.7) if random.random() < 0.15 else None

        for i in range(steps + 1):
            t = i / steps
            # 三次贝塞尔曲线
            x = (1-t)**3 * start_x + 3*(1-t)**2 * t * cp1_x + 3*(1-t)*t**2 * cp2_x + t**3 * target_x
            y = (1-t)**3 * start_y + 3*(1-t)**2 * t * cp1_y + 3*(1-t)*t**2 * cp2_y + t**3 * target_y

            # 添加随机抖动
            if jitter and i > 0 and i < steps:
                x += random.uniform(-2, 2)
                y += random.uniform(-2, 2)

            # 速度变化：中间快，两端慢，加入随机扰动
            speed_factor = 0.015 + 0.035 * math.sin(t * math.pi) + random.uniform(-0.005, 0.005)
            page.mouse.move(x, y)
            time.sleep(max(0.005, speed_factor * random.uniform(0.8, 1.2)))

            # 中途停顿（模拟人看到内容后思考）
            if pause_at is not None and abs(t - pause_at) < 0.02:
                time.sleep(random.uniform(0.3, 0.8))
                pause_at = None

        # 末端微抖动（模拟人手到达目标后的不稳定）
        if random.random() < 0.4:
            for _ in range(random.randint(2, 5)):
                jx = target_x + random.uniform(-3, 3)
                jy = target_y + random.uniform(-3, 3)
                page.mouse.move(jx, jy)
                time.sleep(random.uniform(0.01, 0.03))
            # 最后回到目标点
            page.mouse.move(target_x, target_y)
    except Exception:
        pass


def random_hover(page):
    """随机悬停在页面某个元素上"""
    try:
        selectors = ["a", "div", "section", "img", "span"]
        sel = random.choice(selectors)
        elements = page.query_selector_all(sel)
        if elements:
            elem = random.choice(elements)
            box = elem.bounding_box()
            if box and box["x"] > 0 and box["y"] > 0:
                move_mouse_human(page, box["x"] + box["width"]/2, box["y"] + box["height"]/2)
                human_sleep((0.5, 1.5))
    except Exception:
        pass


def random_click_blank(page):
    """随机点击页面空白处"""
    try:
        x = random.randint(50, VIEWPORT["width"] - 50)
        y = random.randint(50, VIEWPORT["height"] - 50)
        move_mouse_human(page, x, y)
        page.mouse.click(x, y)
        human_sleep((0.3, 0.8))
    except Exception:
        pass


def random_keyboard_activity(page):
    """随机键盘活动：方向键、Esc、空格等，模拟真实用户浏览习惯"""
    try:
        actions = [
            lambda: page.keyboard.press("ArrowDown"),
            lambda: page.keyboard.press("ArrowUp"),
            lambda: page.keyboard.press("ArrowRight"),
            lambda: page.keyboard.press("ArrowLeft"),
            lambda: page.keyboard.press(" "),  # 空格翻页
            lambda: page.keyboard.press("Home"),
            lambda: page.keyboard.press("End"),
        ]
        # 30%概率不做任何键盘活动
        if random.random() < 0.3:
            return
        action = random.choice(actions)
        action()
        time.sleep(random.uniform(0.1, 0.4))
        # 偶尔连续按2-3次
        if random.random() < 0.3:
            for _ in range(random.randint(1, 2)):
                random.choice(actions)()
                time.sleep(random.uniform(0.05, 0.15))
    except Exception:
        pass


def human_scroll(page, direction="down"):
    """人类化滚动：速度变化+阅读停顿+偶尔回滚+键盘滚动混合"""
    try:
        # 25%概率用键盘滚动（PageDown/ArrowDown），更像真实用户
        if random.random() < 0.25 and direction == "down":
            if random.random() < 0.6:
                page.keyboard.press("PageDown")
            else:
                for _ in range(random.randint(3, 6)):
                    page.keyboard.press("ArrowDown")
                    time.sleep(random.uniform(0.05, 0.12))
            time.sleep(random.uniform(0.3, 0.8))
            # 20%概率阅读停顿
            if random.random() < 0.2:
                human_sleep(DELAY_READ_PAUSE)
            return

        if direction == "down":
            scroll_amount = random.randint(400, 900)
        else:
            scroll_amount = -random.randint(150, 350)

        # 分多次滚动，速度先快后慢
        steps = random.randint(4, 10)
        remaining = scroll_amount
        for i in range(steps):
            if i == steps - 1:
                step = remaining
            else:
                # 先快后慢
                ratio = 1.0 - (i / steps) * 0.6
                step = int(remaining * ratio / (steps - i))
                remaining -= step
            page.mouse.wheel(0, step)
            time.sleep(random.uniform(0.03, 0.1))

        # 25%概率阅读停顿
        if random.random() < 0.25:
            human_sleep(DELAY_READ_PAUSE)

        # 15%概率向上回滚
        if random.random() < 0.15 and direction == "down":
            time.sleep(random.uniform(0.3, 0.8))
            back_amount = random.randint(80, 200)
            page.mouse.wheel(0, -back_amount)
            time.sleep(random.uniform(0.2, 0.5))
            # 回滚后再向下滚一点（模拟找位置）
            if random.random() < 0.5:
                time.sleep(random.uniform(0.2, 0.4))
                page.mouse.wheel(0, random.randint(30, 80))
                time.sleep(random.uniform(0.1, 0.3))
    except Exception:
        pass


def random_interact_with_note(page):
    """随机点开一个帖子详情，浏览+评论框输入但不发送，然后返回。增加真实用户行为特征。"""
    try:
        notes = page.query_selector_all("a[href*='/search_result/']")
        if not notes:
            return False

        # 只选有标题文本的帖子
        target_notes = [n for n in notes if n.inner_text().strip()]
        if not target_notes:
            return False

        note = random.choice(target_notes)
        title = note.inner_text().strip()[:30]
        print(f"  随机浏览帖子：{title}...")

        # 移动鼠标到帖子并点击
        box = note.bounding_box()
        if box and box["y"] < VIEWPORT["height"] - 100:
            move_mouse_human(page, box["x"] + box["width"]/2, box["y"] + box["height"]/2)
            human_sleep((0.3, 0.7))
        note.click()
        human_sleep(DELAY_PAGE_LOAD)

        # 等待帖子详情加载
        try:
            page.wait_for_selector("div.note-content, div.detail-container, textarea, input[placeholder*='评论'], input[placeholder*='说点什么']", timeout=12000)
        except Exception:
            pass
        human_sleep((1.0, 2.5))

        # 随机移动鼠标
        move_mouse_human(page)
        human_sleep((0.3, 0.8))

        # 滚动浏览正文和评论（1-3次）
        scroll_times = random.randint(1, 3)
        for _ in range(scroll_times):
            human_scroll(page, "down")
            human_sleep((0.8, 1.8))

        # 60%概率进行评论输入（输入后全选删除，绝不发送）
        if random.random() < 0.6:
            try:
                comment_input = page.query_selector(
                    "textarea, input[placeholder*='评论'], input[placeholder*='说点什么'], div[contenteditable='true']"
                )
                if comment_input:
                    box = comment_input.bounding_box()
                    if box:
                        move_mouse_human(page, box["x"] + box["width"]/2, box["y"] + box["height"]/2)
                        human_sleep((0.3, 0.6))

                    comment_input.click()
                    human_sleep((0.3, 0.7))

                    # 随机输入几个字（常见评论词或无意义字符）
                    random_texts = [
                        "mark", "学习了", "谢谢分享", "666", "收藏了",
                        "有用", "感谢", "m", "1", "看看", "不错"
                    ]
                    text = random.choice(random_texts)
                    for char in text:
                        comment_input.type(char, delay=random.randint(80, 220))
                        # 偶尔停顿思考
                        if random.random() < 0.1:
                            human_sleep((0.2, 0.5))

                    human_sleep((0.5, 1.3))

                    # 输入完成后停顿几秒再删（模拟真实用户分心/思考后再回来删字）
                    human_sleep((3.0, 7.0))

                    # 慢慢逐字删除（模拟人手按退格删字，而非 Ctrl+A 一次性清空）
                    for _ch_i in range(len(text)):
                        comment_input.press("Backspace")
                        time.sleep(random.uniform(0.08, 0.3))
                        # 偶尔停顿（像人在犹豫）
                        if random.random() < 0.1:
                            time.sleep(random.uniform(0.4, 1.2))
                    human_sleep((0.6, 1.5))

                    # 点击空白处失焦
                    random_click_blank(page)
                    print("  评论框已输入并逐字清空（未发送）")
            except Exception:
                pass

        human_sleep((0.5, 1.5))

        # 返回搜索结果页
        page.go_back()
        human_sleep(DELAY_AFTER_SEARCH)

        # 等待搜索结果重新加载
        try:
            page.wait_for_selector("a[href*='/search_result/']", timeout=10000)
        except Exception:
            pass
        human_sleep((0.5, 1.2))
        move_mouse_human(page)
        return True
    except Exception:
        try:
            page.go_back()
            human_sleep((1.0, 2.0))
        except Exception:
            pass
        return False


def human_type(element, text):
    """人类化打字：随机间隔+偶尔删除重输"""
    try:
        element.click()
        human_sleep((0.2, 0.5))
        for i, char in enumerate(text):
            element.type(char, delay=random.randint(60, 180))
            # 5%概率停顿思考
            if random.random() < 0.05 and i < len(text) - 1:
                human_sleep((0.3, 0.8))
            # 2%概率打错删除重输
            if random.random() < 0.02 and i > 0:
                element.press("Backspace")
                time.sleep(random.uniform(0.1, 0.3))
                element.type(char, delay=random.randint(80, 150))
    except Exception:
        element.fill(text)


def clean_author_name(text):
    if not text:
        return ""
    lines = text.strip().split("\n")
    return lines[0].strip() if lines else text.strip()


def is_blacklisted(author_name, author_id=""):
    """检查作者是否在账号黑名单中。
    支持两种匹配：按昵称（BLACKLIST_AUTHORS）精确匹配，或按小红书号（BLACKLIST_IDS）精确匹配。
    均不区分大小写、去除首尾空格。"""
    if author_name:
        name = author_name.strip().lower()
        if any(name == b.strip().lower() for b in BLACKLIST_AUTHORS if b.strip()):
            return True
    if author_id:
        aid = author_id.strip()
        if any(aid == i.strip() for i in BLACKLIST_IDS if i.strip()):
            return True
    return False


def is_irrelevant_content(title):
    """检查帖子标题是否包含不相关关键词黑名单（不区分大小写）。
    命中则返回 True，表示该帖子应被跳过。只检查标题，避免正文误杀。"""
    if not title or not IRRELEVANT_KEYWORDS:
        return False
    title_lower = title.strip().lower()
    return any(kw.strip().lower() in title_lower for kw in IRRELEVANT_KEYWORDS if kw.strip())


_PUB_TIME_PATTERNS = [
    r"刚刚",
    r"\d+\s*分钟前",
    r"\d+\s*小时前",
    r"\d+\s*天前",
    r"昨天",
    r"前天",
    r"\d{4}[-/年]\d{1,2}[-/月]\d{1,2}日?",
    r"\d{1,2}[-/月]\d{1,2}日?",
]


def extract_publish_time(text):
    """从整段文本中用正则提取发布时间，不依赖换行行号结构（原按行取第2行过于脆弱）。"""
    if not text:
        return ""
    for pat in _PUB_TIME_PATTERNS:
        m = re.search(pat, text)
        if m:
            return m.group(0).strip()
    return ""


def parse_xhs_time(time_str):
    if not time_str:
        return None
    now = datetime.now()
    time_str = time_str.strip()
    if time_str == "刚刚":
        return now
    m = re.match(r"(\d+)分钟前", time_str)
    if m:
        return now.replace(microsecond=0) - timedelta(minutes=int(m.group(1)))
    m = re.match(r"(\d+)小时前", time_str)
    if m:
        return now.replace(microsecond=0) - timedelta(hours=int(m.group(1)))
    m = re.match(r"(\d+)天前", time_str)
    if m:
        return now.replace(microsecond=0) - timedelta(days=int(m.group(1)))
    if time_str.startswith("昨天"):
        return now.replace(microsecond=0) - timedelta(days=1)
    if time_str.startswith("前天"):
        return now.replace(microsecond=0) - timedelta(days=2)
    m = re.match(r"(\d{4})年(\d{1,2})月(\d{1,2})日?", time_str)
    if m:
        try:
            return datetime(int(m.group(1)), int(m.group(2)), int(m.group(3)))
        except ValueError:
            return None
    m = re.match(r"(\d{4})[-/](\d{1,2})[-/](\d{1,2})", time_str)
    if m:
        try:
            return datetime(int(m.group(1)), int(m.group(2)), int(m.group(3)))
        except ValueError:
            return None
    # M月D日 / MM-DD：按今年解析，若晚于当前则视为去年（减少误判）
    m = re.match(r"(\d{1,2})月(\d{1,2})日?", time_str)
    if m:
        try:
            d = datetime(now.year, int(m.group(1)), int(m.group(2)))
            if d > now:
                d = d.replace(year=now.year - 1)
            return d
        except ValueError:
            return None
    m = re.match(r"(\d{1,2})[-/](\d{1,2})", time_str)
    if m:
        try:
            d = datetime(now.year, int(m.group(1)), int(m.group(2)))
            if d > now:
                d = d.replace(year=now.year - 1)
            return d
        except ValueError:
            return None
    return None


def is_within_two_years(time_str):
    """True=两年内保留 / False=超过两年丢弃 / None=无法解析（由调用方决定，默认保留但不静默）。"""
    pub_time = parse_xhs_time(time_str)
    if pub_time is None:
        return None
    two_years_ago = datetime.now() - timedelta(days=730)
    return pub_time >= two_years_ago


def extract_note_id(url):
    match = re.search(r"/search_result/([a-zA-Z0-9]+)", url)
    if match:
        return match.group(1)
    match = re.search(r"/explore/([a-zA-Z0-9]+)", url)
    if match:
        return match.group(1)
    return url


def is_logged_in(page):
    """以'未登录信号'为主判断：登录弹窗→未登录；头像→已登录；顶部登录入口→未登录；搜索框仅作兜底。"""
    try:
        # 1) 出现登录弹窗（二维码/手机号登录）→ 明确未登录
        for sel in [
            "div[class*='login-container']",
            "div[class*='login-modal']",
            "div[class*='login-dialog']",
            "div[class*='login-panel']",
            "div[class*='qrcode']",
        ]:
            if page.query_selector(sel):
                return False

        # 2) 已登录常见头像
        if page.query_selector("div[class*='avatar'], img[class*='avatar']"):
            return True

        # 3) 未登录时顶部常见的显式"登录"入口
        for tag in ("a", "button"):
            try:
                if page.query_selector(f"{tag} >> text=登录"):
                    return False
            except Exception:
                continue

        # 4) 最后兜底：存在搜索框视为可继续操作
        search_input = page.query_selector("input#search-input, input[placeholder*='搜索']")
        return bool(search_input)
    except Exception:
        return False


def goto_with_referer(page, url, timeout=30000):
    """带 Referer 的页面跳转：模拟从当前页"点击"进入目标页，避免 Referer 为空被识别为程序化跳转。

    真实用户从 A 页点击链接进 B 页时，B 页请求必带 Referer=A。
    直接 page.goto() 会让 Referer 为空，服务器会将其判定为脚本直连。
    经实测，route.continue_ 无法覆盖 Referer（Chromium 忽略），必须用 set_extra_http_headers 才能发出。
    """
    referer_url = ""
    try:
        cur = page.url or ""
        if cur and "about:blank" not in cur:
            referer_url = cur
    except Exception:
        pass
    if referer_url:
        try:
            page.set_extra_http_headers({"Referer": referer_url})
        except Exception:
            pass
    try:
        page.goto(url, wait_until="domcontentloaded", timeout=timeout)
    finally:
        # 用完即清，避免影响后续页面的 Referer 语义
        try:
            page.set_extra_http_headers({})
        except Exception:
            pass


def search_keyword(page, keyword):
    """通过首页搜索框搜索，模拟人类操作"""
    try:
        page.goto("https://www.xiaohongshu.com", wait_until="domcontentloaded", timeout=30000)
        human_sleep(DELAY_PAGE_LOAD)

        # 随机交互
        if random.random() < 0.5:
            random_hover(page)
        if random.random() < 0.3:
            random_click_blank(page)

        search_input = page.wait_for_selector("input#search-input, input[placeholder*='搜索']", timeout=10000)
        if search_input:
            box = search_input.bounding_box()
            if box:
                move_mouse_human(page, box["x"] + box["width"]/2, box["y"] + box["height"]/2)
                human_sleep(DELAY_MOUSE_MOVE)
                search_input.click()
                human_sleep((0.3, 0.6))

            search_input.fill("")
            human_type(search_input, keyword)
            human_sleep((0.5, 1.2))
            search_input.press("Enter")
        else:
            search_url = f"https://www.xiaohongshu.com/search_result?keyword={quote(keyword)}"
            goto_with_referer(page, search_url)
    except Exception:
        search_url = f"https://www.xiaohongshu.com/search_result?keyword={quote(keyword)}"
        try:
            goto_with_referer(page, search_url)
        except Exception:
            print(f"  警告：搜索 URL 直连也失败，关键词「{keyword}」将跳过")

    human_sleep(DELAY_AFTER_SEARCH)

    try:
        page.wait_for_selector("a[href*='/search_result/']", timeout=15000)
    except Exception:
        print(f"  警告：等待搜索结果超时")
    human_sleep((1.0, 2.5))

    move_mouse_human(page)
    if random.random() < 0.4:
        random_hover(page)


def extract_note_content_with_disguise(context, note_url):
    """随机打开帖子详情页读正文，带真人浏览伪装。
    强化反检测：超时缩短、风控页检测、分多步先快后慢滚动、随机停留。
    失败或触发风控返回空字符串。"""
    detail_page = None
    try:
        # ===== 修复 Referer 链：模拟从当前搜索结果页"点击"进入详情 =====
        # 真实用户从搜索列表点击卡片进详情，Referer 必为搜索结果页 URL。
        # 直接 goto() 会让 Referer 为空，被服务器识别为程序化跳转。
        # 注意：route.continue_ 覆盖 Referer 无效（Chromium 会忽略），
        # 必须用 set_extra_http_headers 才能真正发出 Referer 头。
        referer_url = ""
        try:
            for pg in reversed(context.pages):
                u = (pg.url or "")
                if u and "about:blank" not in u:
                    referer_url = u
                    break
        except Exception:
            pass

        detail_page = context.new_page()

        if referer_url:
            try:
                detail_page.set_extra_http_headers({"Referer": referer_url})
            except Exception:
                pass

        try:
            detail_page.goto(note_url, wait_until="domcontentloaded", timeout=12000)
        except Exception:
            return ""

        # 风控检测：页面跳转到安全限制/验证页，或正文出现强风控信号时立即退出并暂停
        try:
            cur_url = detail_page.url or ""
            if any(k in cur_url for k in ("website-login", "error_code", "verify")):
                human_sleep((8.0, 15.0))
                return ""
            page_text = detail_page.inner_text("body")[:500]
            if any(kw in page_text for kw in ["访问频繁", "操作频繁", "安全验证", "请完成验证",
                                               "滑动验证", "行为异常", "暂时无法", "触发风控"]):
                human_sleep((8.0, 15.0))
                return ""
        except Exception:
            pass

        # 等待正文加载（最多6秒）
        content_selectors = [
            "div#detail-desc",
            "div.note-content",
            "div[class*='note-content']",
            "div[class*='desc']",
            "span[class*='content']",
            "div.detail-container",
            "div[class*='note-text']",
        ]
        try:
            detail_page.wait_for_selector(", ".join(content_selectors), timeout=6000)
        except Exception:
            pass

        # ===== 伪装：真人阅读行为 =====
        # 1) 阅读停顿（模拟认真看正文）
        human_sleep((2.5, 6.0))
        # 2) 随机滚动 0-2 次（分多步先快后慢，模拟真人滚动惯性）
        scroll_times = random.randint(0, 2)
        for _ in range(scroll_times):
            try:
                total = random.randint(200, 500)
                steps = random.randint(3, 6)
                remaining = total
                for si in range(steps):
                    if si == steps - 1:
                        step = remaining
                    else:
                        ratio = 1.0 - (si / steps) * 0.6
                        step = int(remaining * ratio / (steps - si))
                        remaining -= step
                    detail_page.mouse.wheel(0, step)
                    time.sleep(random.uniform(0.03, 0.08))
                human_sleep((0.5, 1.5))
            except Exception:
                pass
        # 3) 60%概率鼠标随机移动（模拟浏览视线）
        if random.random() < 0.6:
            try:
                move_mouse_human(detail_page)
            except Exception:
                pass

        # 读取正文（取最长的匹配）
        content = ""
        for sel in content_selectors:
            try:
                el = detail_page.query_selector(sel)
                if el:
                    text = el.inner_text().strip()
                    if len(text) > len(content):
                        content = text
                    if len(text) >= 20:
                        break
            except Exception:
                continue
        # 选择器都没找到时，从 meta description 兜底
        if not content:
            try:
                meta = detail_page.query_selector("meta[name='description']")
                if meta:
                    content = meta.get_attribute("content") or ""
            except Exception:
                pass
        if content:
            content = re.sub(r"\n{3,}", "\n\n", content).strip()

        # 关闭前再停顿一下（模拟看完了准备退出）
        human_sleep((0.8, 2.0))
        return content

    except Exception:
        return ""
    finally:
        if detail_page:
            try:
                detail_page.close()
            except Exception:
                pass


def scroll_and_collect(page, context, keyword, max_count):
    collected = []
    seen_ids = set()
    last_count = 0
    no_new_count = 0
    unknown_kept = 0  # 发布时间无法解析但默认保留的条数（用于非静默提示）

    for scroll_idx in range(MAX_SCROLLS):
        notes = page.query_selector_all("a[href*='/search_result/']")

        for note_a in notes:
            try:
                href = note_a.get_attribute("href") or ""
                full_url = urljoin("https://www.xiaohongshu.com", href)
                note_id = extract_note_id(full_url)

                if note_id in seen_ids:
                    continue

                title = note_a.inner_text().strip()
                if not title:
                    continue

                card = note_a.evaluate_handle("el => el.closest('section, div.note-item, .feeds-page .note-item')")
                author_name = ""
                author_id = ""
                publish_time = ""
                card_el = card.as_element() if card else None
                if card_el:
                    # 1) 优先从卡片内独立的时间元素定位发布时间
                    for tsel in TIME_SELECTORS:
                        try:
                            tel = card_el.query_selector(tsel)
                            if tel:
                                _t = tel.inner_text().strip()
                                if _t:
                                    publish_time = _t
                                    break
                        except Exception:
                            continue
                    # 2) 时间元素定位失败时，从卡片整体文本用正则兜底提取
                    if not publish_time:
                        publish_time = extract_publish_time(card_el.inner_text())
                    # 3) 作者名从用户链接取（取首行）
                    author_a = card_el.query_selector("a[href*='/user/profile/']")
                    if author_a:
                        author_name = clean_author_name(author_a.inner_text())
                        _ahref = author_a.get_attribute("href") or ""
                        _m = re.search(r"/user/profile/([a-zA-Z0-9]+)", _ahref)
                        if _m:
                            author_id = _m.group(1)

                # 账号黑名单过滤：跳过黑名单账号（昵称或小红书号命中）发布的帖子
                if is_blacklisted(author_name, author_id):
                    print(f"    跳过黑名单账号：{author_name}")
                    continue

                if FILTER_WITHIN_TWO_YEARS:
                    within = is_within_two_years(publish_time)
                    if within is False:
                        continue
                    if within is None and unknown_kept < 3:
                        unknown_kept += 1
                        print(f"    提示：标题「{title[:20]}」发布时间[{publish_time or '未知'}]无法解析，默认保留")

                # 内容不相关筛选：标题包含黑名单关键词的帖子跳过（避免为不相关内容浪费时间）
                if is_irrelevant_content(title):
                    print(f"    跳过不相关内容：{title[:40]}...")
                    continue

                # 先标记已处理（防止详情页打开异常导致重复打开同一个帖子）
                seen_ids.add(note_id)

                # 随机打开详情页（30%概率，模拟真人不会每条都点进去看）
                note_content = ""
                if random.random() < 0.3:
                    note_content = extract_note_content_with_disguise(context, full_url)
                    # 点开看完后随机停顿（模拟回味/思考）
                    human_sleep((1.5, 4.0))
                else:
                    # 不点开时也停顿一下（模拟快速划过）
                    human_sleep((0.8, 2.0))

                collected.append({
                    "搜索关键词": keyword,
                    "账号名字": author_name,
                    "小红书号": author_id,
                    "帖子标题": title,
                    "帖子正文": note_content,
                    "帖子链接": full_url,
                    "发布时间": publish_time,
                })

                # 实时落盘：采到一条立即写入实时文件，防止中途退出丢失已采集数据
                _live_save(collected)

                if note_content:
                    content_info = f"已点开·正文{len(note_content)}字"
                else:
                    content_info = "未点开"
                print(f"    采集到：{title}... ({content_info})")

                if len(collected) >= max_count:
                    return collected

            except Exception:
                continue

        print(f"  第{scroll_idx + 1}次滚动，已采集 {len(collected)} 条")

        if len(collected) == last_count:
            no_new_count += 1
            if no_new_count >= 3:
                print("  连续3次无新内容，停止滚动")
                break
        else:
            no_new_count = 0
        last_count = len(collected)

        human_scroll(page, "down")
        human_sleep(DELAY_AFTER_SCROLL)

        # 随机交互
        if random.random() < 0.5:
            move_mouse_human(page)
        if random.random() < 0.3:
            random_hover(page)
        if random.random() < 0.15:
            random_click_blank(page)
        if random.random() < 0.2:
            random_keyboard_activity(page)

        # 12%概率随机点开帖子详情交互（浏览+评论输入但不发送），至少采集3条后才触发
        if random.random() < 0.12 and len(collected) >= 3:
            random_interact_with_note(page)

    return collected


def save_to_excel(data, filepath):
    wb = Workbook()
    ws = wb.active
    ws.title = "小红书采集结果"

    headers = ["序号", "搜索关键词", "账号名字", "小红书号", "帖子标题", "帖子正文", "帖子链接", "发布时间"]
    col_widths = [6, 16, 20, 22, 40, 80, 60, 12]

    header_font = Font(name="微软雅黑", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="FF2442", end_color="FF2442", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9"),
    )

    for col_idx, (header, width) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border
        ws.column_dimensions[cell.column_letter].width = width

    data_font = Font(name="微软雅黑", size=10)
    data_align = Alignment(horizontal="left", vertical="top", wrap_text=True)
    for row_idx, item in enumerate(data, 2):
        values = [
            row_idx - 1,
            item["搜索关键词"],
            item["账号名字"],
            item.get("小红书号", ""),
            item["帖子标题"],
            item.get("帖子正文", ""),
            item["帖子链接"],
            item.get("发布时间", ""),
        ]
        for col_idx, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font = data_font
            cell.alignment = data_align
            cell.border = thin_border

    ws.row_dimensions[1].height = 30
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    wb.save(filepath)
    return filepath


def _live_save(rows):
    """实时落盘：把当前已有结果写入实时文件（固定名），失败不阻断主流程。"""
    try:
        save_to_excel(rows, LIVE_OUTPUT_FILE)
    except Exception:
        pass


def _run_keywords(page, context):
    # 关键词顺序每次运行随机打乱：打散跨会话行为画像，避免固定扫描模式
    kws = list(KEYWORDS)
    random.shuffle(kws)
    for idx, kw in enumerate(kws):
        print(f"【{kw}】正在搜索...")
        try:
            # 每词采集量加随机浮动（下限1条，上限 MAX_PER_KEYWORD），避免每次都是固定10条
            cap = max(1, int(MAX_PER_KEYWORD * random.uniform(0.5, 1.0)))
            results = scroll_and_collect(page, context, kw, cap)

            new_count = 0
            for item in results:
                nid = extract_note_id(item["帖子链接"])
                if nid not in _seen_note_ids:
                    _seen_note_ids.add(nid)
                    _all_rows.append(item)
                    new_count += 1

            print(f"【{kw}】采集完成，本关键词 {len(results)} 条（目标{cap}），去重后新增 {new_count} 条\n")
        except Exception as e:
            print(f"【{kw}】采集出错：{e}\n")
            continue

        # 单次总量上限：达到后提前结束，防止单 IP 一次抓太多
        if MAX_TOTAL_NOTES and len(_all_rows) >= MAX_TOTAL_NOTES:
            print(f"\n已达单次总量上限 {MAX_TOTAL_NOTES} 条，提前结束采集。")
            break

        if idx < len(kws) - 1:
            print(f"  等待 {int(DELAY_BETWEEN_KEYWORDS[0])}-{int(DELAY_BETWEEN_KEYWORDS[1])} 秒后继续...")
            # 等待期间模拟人类浏览
            wait_end = time.time() + _skewed_random(*DELAY_BETWEEN_KEYWORDS)
            while time.time() < wait_end:
                if random.random() < 0.6:
                    move_mouse_human(page)
                if random.random() < 0.3:
                    random_hover(page)
                if random.random() < 0.2:
                    human_scroll(page, random.choice(["down", "up"]))
                if random.random() < 0.1:
                    random_click_blank(page)
                human_sleep((1.5, 3.0))


def main():
    print("=" * 60)
    print("小红书关键词帖子采集工具（反检测加固版 v1）")
    print("=" * 60)
    print(f"关键词列表：{KEYWORDS}（顺序每次随机打乱）")
    print(f"每个关键词最多采集：{MAX_PER_KEYWORD} 条（实际随机浮动）")
    print(f"单次总采集上限：{MAX_TOTAL_NOTES} 条")
    print(f"代理：{PROXY_URL or '直连（本机 IP）'}")
    print(f"两年内过滤：{'开启' if FILTER_WITHIN_TWO_YEARS else '关闭'}")
    print(f"输出文件：{OUTPUT_FILE}")
    print("=" * 60)

    with sync_playwright() as p:
        launch_args = [
            "--disable-blink-features=AutomationControlled",
            "--disable-infobars",
            "--window-size=1920,1080",
            "--lang=zh-CN",
            "--no-first-run",
            "--disable-notifications",
            "--exclude-switches=enable-automation",
        ]
        # --no-sandbox / --disable-dev-shm-usage 仅 Linux/容器环境需要，Windows 下无用
        if not sys.platform.startswith("win"):
            launch_args += ["--no-sandbox", "--disable-dev-shm-usage"]

        context = None
        used_channel = None
        # 使用持久化用户目录：登录态可跨运行保留，避免每次重新扫码（也减少重复登录触发风控）
        for channel in ["chrome", "msedge"]:
            try:
                context = p.chromium.launch_persistent_context(
                    user_data_dir=PROFILE_DIR,
                    headless=False,
                    channel=channel,
                    args=launch_args,
                    viewport=VIEWPORT,
                    locale="zh-CN",
                    timezone_id="Asia/Shanghai",
                    permissions=["notifications"],
                    geolocation={"latitude": 29.5630, "longitude": 106.5516},
                    color_scheme="light",
                    ignore_default_args=["--enable-automation"],
                    **({"proxy": {"server": PROXY_URL}} if PROXY_URL else {}),
                )
                used_channel = channel
                print(f"使用浏览器：{channel}")
                break
            except Exception:
                continue
        if not context:
            # 本机无 Chrome/Edge 时，尝试 Playwright 自带 Chromium
            print("\n未找到系统 Chrome/Edge，尝试 Playwright 自带 Chromium...")
            try:
                context = p.chromium.launch_persistent_context(
                    user_data_dir=PROFILE_DIR,
                    headless=False,
                    args=launch_args,
                    viewport=VIEWPORT,
                    locale="zh-CN",
                    timezone_id="Asia/Shanghai",
                    permissions=["notifications"],
                    geolocation={"latitude": 29.5630, "longitude": 106.5516},
                    color_scheme="light",
                    ignore_default_args=["--enable-automation"],
                    **({"proxy": {"server": PROXY_URL}} if PROXY_URL else {}),
                )
                used_channel = "playwright-chromium"
            except Exception:
                context = None
        if not context:
            raise RuntimeError(
                "未找到可用浏览器。请安装 Google Chrome 或 Microsoft Edge，"
                "或执行 'playwright install chromium' 安装 Playwright 自带浏览器后重试。"
            )

        # 资源拦截：禁用视频/字体（保留图片，方便观察筛选出的帖子卡片），屏蔽统计/反爬域名
        def _resource_route(route):
            try:
                req = route.request
                rtype = req.resource_type
                url = req.url
                # 修复：拦截 favicon.ico —— 真实浏览器 favicon 有缓存，不会每次跳转都重新请求。
                # 之前每次 goto 后紧跟 0.02s 的 /favicon.ico 请求，形成"页面+favicon 双峰"时序特征。
                if "favicon" in url.lower() or url.rstrip("/").endswith(".ico"):
                    route.abort()
                    return
                if rtype in ("media", "font"):
                    route.abort()
                    return
                _blocked = ("hm.baidu.com", "sensorsdata.cn", "sensorsdata.com",
                            "log.snssdk.com", "xlog.snssdk.com", "applog.snssdk.com",
                            "byteoversea.com", "beacon.qq.com", "google-analytics.com",
                            "googletagmanager.com", "doubleclick.net", "scorecardresearch.com")
                if any(d in url for d in _blocked):
                    route.abort()
                    return
                route.continue_()
            except Exception:
                try:
                    route.continue_()
                except Exception:
                    pass
        context.route("**/*", _resource_route)

        page = context.new_page()

        # 读取真实 User-Agent 与硬件/平台属性，动态生成注入脚本
        # 修复硬编码硬件值(如固定16核/16G/RTX3060)与真机不一致反而暴露自动化特征的问题
        _NAV_JS = (
            "() => {"
            "  const c = document.createElement('canvas');"
            "  const gl = c.getContext('webgl') || c.getContext('experimental-webgl');"
            "  let v = '', r = '';"
            "  try {"
            "    if (gl) {"
            "      const dbg = gl.getExtension('WEBGL_debug_renderer_info');"
            "      if (dbg) {"
            "        v = gl.getParameter(dbg.UNMASKED_VENDOR_WEBGL) || '';"
            "        r = gl.getParameter(dbg.UNMASKED_RENDERER_WEBGL) || '';"
            "      }"
            "    }"
            "  } catch(e){}"
            "  const pl = [];"
            "  try { for (let i = 0; i < navigator.plugins.length; i++) {"
            "    const p = navigator.plugins[i];"
            "    pl.push({name: p.name, filename: p.filename, description: p.description});"
            "  } } catch(e){}"
            "  const mm = [];"
            "  try { for (let j = 0; j < navigator.mimeTypes.length; j++) {"
            "    const m = navigator.mimeTypes[j];"
            "    mm.push({type: m.type, suffixes: m.suffixes, description: m.description});"
            "  } } catch(e){}"
            "  return {"
            "    hw: navigator.hardwareConcurrency || 8,"
            "    mem: navigator.deviceMemory || 8,"
            "    mtp: navigator.maxTouchPoints || 0,"
            "    platform: navigator.platform || 'Win32',"
            "    vendor: navigator.vendor || 'Google Inc.',"
            "    webgl_vendor: v,"
            "    webgl_renderer: r,"
            "    plugins: pl,"
            "    mimes: mm"
            "  };"
            "}"
        )
        page.goto("about:blank")
        real_ua = page.evaluate("navigator.userAgent") or USER_AGENT
        try:
            real_nav = page.evaluate(_NAV_JS)
        except Exception:
            real_nav = {}
        context.add_init_script(build_stealth_script(real_ua, real_nav))

        # UA-CH 预热：主动请求一次 high-entropy 值，促使浏览器在后续请求自然携带完整 Sec-CH-UA-* 头
        try:
            page.evaluate(
                "navigator.userAgentData && navigator.userAgentData.getHighEntropyValues ? "
                "navigator.userAgentData.getHighEntropyValues("
                "['architecture','bitness','platformVersion','fullVersionList']).catch(function(){}) : Promise.resolve()"
            )
        except Exception:
            pass

        # 请求头交给真实浏览器原生发送，不再强制覆盖（避免 UA / Client-Hints 与真实版本冲突）

        print("\n正在打开小红书...")
        page.goto("https://www.xiaohongshu.com", wait_until="domcontentloaded", timeout=30000)
        human_sleep(DELAY_PAGE_LOAD)

        # 初始随机交互
        move_mouse_human(page)
        human_sleep((1.0, 2.5))
        if random.random() < 0.5:
            random_hover(page)
        if random.random() < 0.3:
            random_click_blank(page)

        if not is_logged_in(page):
            print("\n" + "!" * 50)
            print("检测到未登录！请在弹出的浏览器中手动完成登录")
            print("脚本会自动检测登录状态，登录成功后自动继续...")
            print("!" * 50)
            waited = 0
            while waited < 300:
                time.sleep(3)
                waited += 3
                try:
                    if is_logged_in(page):
                        print(f"\n检测到登录成功（等待约 {waited} 秒），继续执行...\n")
                        break
                except Exception:
                    pass
            else:
                print("\n等待登录超时，脚本退出。")
                context.close()
                return
            human_sleep(DELAY_PAGE_LOAD)

        print("登录状态确认，开始采集...\n")

        try:
            _run_keywords(page, context)
        except KeyboardInterrupt:
            print("\n\n收到中断信号（Ctrl+C），正在保存已采集的数据...")
            context.close()
            if _all_rows:
                save_to_excel(_all_rows, OUTPUT_FILE)
                print(f"已保存 {len(_all_rows)} 条到：{OUTPUT_FILE}")
            else:
                print("本次尚未采集到任何数据。")
            return
        except Exception as e:
            print(f"\n\n采集过程发生异常：{e}")
            context.close()
            if _all_rows:
                save_to_excel(_all_rows, OUTPUT_FILE)
                print(f"已保存 {len(_all_rows)} 条到：{OUTPUT_FILE}")
            else:
                print("本次尚未采集到任何数据。")
            return

        context.close()

    if _all_rows:
        save_path = save_to_excel(_all_rows, OUTPUT_FILE)
        print(f"\n全部完成！共采集 {len(_all_rows)} 条不重复帖子")
        print(f"Excel 已保存至：{save_path}")
    else:
        print("\n未采集到任何数据，请检查网络和登录状态。")


if __name__ == "__main__":
    main()
